"""Sync Status + Date Applied from Excel back into SQLite."""
from __future__ import annotations
import argparse
import os
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", required=True)
    ap.add_argument("--excel", required=True)
    args = ap.parse_args()

    excel_path = Path(os.path.expandvars(args.excel))
    if not excel_path.exists():
        print(f"[ERR] Excel not found: {excel_path}", file=sys.stderr)
        return 1

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["Applications"]

    # Expected headers (from finalize.py)
    headers = [c.value for c in ws[1]]
    try:
        i_company = headers.index("Company")
        i_title = headers.index("Job Title")
        i_applied = headers.index("Date Applied")
        i_status = headers.index("Status")
        i_notes = headers.index("Notes")
    except ValueError as e:
        print(f"[ERR] missing expected header: {e}", file=sys.stderr)
        return 2

    conn = sqlite3.connect(args.db)
    updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        company = row[i_company]
        title = row[i_title]
        status = row[i_status]
        applied = row[i_applied]
        notes = row[i_notes]
        if not company or not title:
            continue
        cur = conn.execute(
            "UPDATE jobs SET status = COALESCE(?, status), applied_at = ?, notes = COALESCE(?, notes) "
            "WHERE LOWER(company) = LOWER(?) AND LOWER(title) = LOWER(?)",
            (
                (status or "").lower() if status else None,
                str(applied) if applied else None,
                notes,
                str(company),
                str(title),
            ),
        )
        updated += cur.rowcount
    conn.commit()
    conn.close()
    print(f"[OK] synced {updated} rows from Excel to SQLite", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
