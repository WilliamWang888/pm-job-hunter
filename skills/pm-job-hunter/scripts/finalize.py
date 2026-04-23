"""Finalize: persist LLM scores, export Excel tracker, emit digest for email."""
from __future__ import annotations
import argparse
import json
import os
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


EXCEL_HEADERS = [
    "Date Matched", "Company", "Job Title", "Location", "Source",
    "Match Score", "Job URL", "Apply URL", "Job Description",
    "Date Applied", "Status", "Notes",
]


def is_file_locked(path: Path) -> bool:
    """Heuristic lock check for Excel file on Windows."""
    if not path.exists():
        return False
    # Check for Excel owner file (~$filename)
    owner = path.parent / f"~${path.name}"
    if owner.exists():
        return True
    try:
        with open(path, "a+b"):
            return False
    except (PermissionError, OSError):
        return True


def refresh_excel(db_path: Path, excel_path: Path) -> tuple[bool, str]:
    """Rewrite the Excel export from SQLite. Returns (success, message)."""
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    if is_file_locked(excel_path):
        return False, "Excel file is open — skipped refresh; SQLite is still updated."

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        """SELECT first_seen_at, company, title, location, source, score,
                  url, apply_url, description, applied_at, status, notes
           FROM jobs ORDER BY first_seen_at DESC"""
    ).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for col_idx, h in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    # Rows
    for r_idx, row in enumerate(rows, start=2):
        # Truncate description for Excel readability
        desc = row["description"] or ""
        if len(desc) > 32000:
            desc = desc[:32000] + "..."
        values = [
            (row["first_seen_at"] or "")[:10],
            row["company"],
            row["title"],
            row["location"],
            row["source"],
            row["score"],
            row["url"],
            row["apply_url"],
            desc,
            (row["applied_at"] or "")[:10] if row["applied_at"] else "",
            (row["status"] or "matched").capitalize(),
            row["notes"] or "",
        ]
        for c_idx, v in enumerate(values, start=1):
            ws.cell(row=r_idx, column=c_idx, value=v)

    # Column widths
    widths = [13, 22, 40, 28, 12, 8, 45, 45, 60, 13, 12, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    try:
        wb.save(excel_path)
    except PermissionError:
        return False, "Excel file is open — skipped refresh; SQLite is still updated."
    return True, f"Excel refreshed: {len(rows)} total rows."


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--ranked", required=True, help="ranked.json from LLM step")
    ap.add_argument("--db", required=True)
    ap.add_argument("--excel", required=True)
    ap.add_argument("--out", required=True, help="digest.json for email composer")
    args = ap.parse_args()

    with open(args.ranked, "r", encoding="utf-8") as f:
        ranked = json.load(f)

    db_path = Path(args.db)
    excel_path = Path(os.path.expandvars(args.excel))

    # Persist scores to SQLite
    conn = sqlite3.connect(db_path)
    updated = 0
    for item in ranked:
        jid = item.get("id")
        score = item.get("score")
        why = item.get("why_match", "")
        if not jid:
            continue
        conn.execute("UPDATE jobs SET score = ?, why_match = ? WHERE id = ?", (score, why, jid))
        updated += conn.total_changes
    conn.commit()

    # Fetch full details for digest
    conn.row_factory = sqlite3.Row
    digest_items: list[dict[str, Any]] = []
    for item in sorted(ranked, key=lambda x: x.get("score", 0), reverse=True):
        jid = item.get("id")
        row = conn.execute(
            "SELECT company, title, location, url, apply_url FROM jobs WHERE id = ?", (jid,)
        ).fetchone()
        if not row:
            continue
        digest_items.append({
            "id": jid,
            "company": row["company"],
            "title": row["title"],
            "location": row["location"],
            "url": row["url"],
            "apply_url": row["apply_url"],
            "score": item.get("score"),
            "why_match": item.get("why_match", ""),
        })
    conn.close()

    # Excel refresh
    excel_ok, excel_msg = refresh_excel(db_path, excel_path)
    print(f"[excel] {excel_msg}", file=sys.stderr)

    digest = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "date_label": datetime.now().strftime("%B %d, %Y"),
        "items": digest_items[:15],
        "excel": {
            "path": str(excel_path),
            "refreshed": excel_ok,
            "message": excel_msg,
        },
    }
    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(digest, f, ensure_ascii=False, indent=2)
    print(f"[OK] wrote digest for {len(digest['items'])} items to {args.out}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
